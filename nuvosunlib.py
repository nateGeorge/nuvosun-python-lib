'''
to do:
save XRF data in a stashed file
'''

from openpyxl import load_workbook
from scipy.integrate import simps
import csv, os, operator, datetime, shutil, errno, time, re, distutils.dir_util, glob, pickle
import sqlite3 as sql
import numpy as np
from dateutil.parser import parse as dateParser
from operator import itemgetter

def getReliRunsDict():
    # gets a dictionary of good/bad reliability runs for SC
    goodReliRuns = ['286', '307', '328', '379', '394', '405', '408']
    goodReliTTF = [2100,2200,1433,1913,1254,1867,2000]
    ReliRunsDict = {}
    counter = 0
    for run in goodReliRuns:
        ReliRunsDict[run] = {}
        ReliRunsDict[run]['bake time'] = '12h'
        ReliRunsDict[run]['SC result'] = 'good'
        ReliRunsDict[run]['SC TTF'] = goodReliTTF[counter]
        counter += 1

    ReliRunsDict['406'] = {}
    ReliRunsDict['406']['bake time'] = '72h'
    ReliRunsDict[run]['SC result'] = 'good'
    ReliRunsDict[run]['SC TTF'] = 1700

    badReliRuns = ['304', '327', '384', '406', '426', '428', '430', '431']
    badReliTTF = [352, 212, 152, 298, 118, 110, 98, 124, 150]
    counter = 0
    for run in badReliRuns:
        ReliRunsDict[run] = {}
        ReliRunsDict[run]['bake time'] = '12h'
        ReliRunsDict[run]['SC result'] = 'bad'
        ReliRunsDict[run]['SC TTF'] = badReliTTF[counter]
        counter += 1

    return ReliRunsDict

def getLatestEffFile():
    effFileDir = 'Y:/Nate/all eff data/'
    TEMPeffFiles = os.listdir(effFileDir)
    effFiles = []
    for file in TEMPeffFiles:
        if os.path.isfile(effFileDir + file):
            effFiles.append(file)
    sortedEffFiles = sorted(effFiles, key=lambda x: os.stat(os.path.join(effFileDir, x)).st_mtime)
    latestEffFile = sortedEffFiles[-1]
    return effFileDir + latestEffFile

def get_addenda_eff_files():
    """Returns an array of efficiency addenda files for updating efficiency data.
    """
    addendaFolder = 'Y:/Nate/all eff data/addenda/'
    fileList = os.listdir(addendaFolder)
    addendaList = []
    for thing in fileList:
        thing = addendaFolder + thing
        if os.path.isfile(thing):
            addendaList.append(thing)
    return addendaList
    
def import_eff_file(effFile = getLatestEffFile(), effCutoff = 0, stashFile = True, substrateRange = [0,1000]):
    # imports efficiency file into dictionary (which is returned), with primary keys as substrates, secondary keys as web IDs
    # crapSubstrateLabels = ['0','110110','GLOBAL','NA','REF01','SPECIAL NEW','SPECIAL OLD','SPECIAL']
    # effCutoff allows you to exclude data with efficiency below effCutoff
    # stash_file will save the effData dict in a pickle file so you don't have to process if the efficiency file isn't new
    
    effData = {}
    colsToImport = ['DW','CW','BC Run','BE Run','SE Run','PC Run','CDS Run','TCO Run',
        'PC Tool','Baked','Cell Eff Avg',
        'Cell Voc Avg','Cell Jsc Avg','Cell FF Avg','Cell Rs Avg','Cell Rsh Avg', 'BE Recipe', 
        'BC Recipe', 'PC Recipe', 'Se Recipe', 'TCO Recipe', 'Cds Recipe', 'Substrate Lot', 'Cell ID'] # 'BC Tool','BE Tool','DateTested','Se Tool', 'Cds Tool','TCO Tool',
    realTypes = ['DW','CW','Cell Eff Avg',
        'Cell Voc Avg','Cell Jsc Avg','Cell FF Avg','Cell Rs Avg','Cell Rsh Avg']
    # other columns you could import: 'Bake Duration','LightSoak'
    
    if stashFile:
        effDBFile = 'Y:/Nate/all eff data/stash/effData.db'
        effDBConn = sql.connect(effDBFile)
        effDBConn.text_factory = str
        cursor = effDBConn.cursor()
        logFile = 'Y:/Nate/all eff data/stash/effLog.pkl'
    
    addendaList = get_addenda_eff_files()
    
    filesToProcess = []
    filesProcessed = []
    
    if os.path.isfile(logFile) and stashFile:
        filesProcessed = pickle.load(open(logFile))
        for file in (addendaList + [effFile]):
            if file not in filesProcessed:
                print 'haven\'t processed', file
                filesToProcess.append(file)
        if len(filesToProcess) == 0:
            print 'loading data from saved db'
            startTime = datetime.datetime.now()
            cursor.execute("SELECT * FROM effTable WHERE substrate BETWEEN {min} AND {max} AND \"Cell Eff Avg\" > {effCut};".format(min = substrateRange[0], max = substrateRange[1], effCut = effCutoff))
            while True:
                dataRow = cursor.fetchone()
                if dataRow == None:
                    break
                effData.setdefault(dataRow[0],{})
                effData[dataRow[0]].setdefault(dataRow[1],{})
                colCounter = 2
                for col in colsToImport:
                    effData[dataRow[0]][dataRow[1]].setdefault(col,[]).append(dataRow[colCounter])
                    colCounter += 1
            print 'loaded eff data from saved file'
            print 'took', (datetime.datetime.now()-startTime).total_seconds(), 'seconds'
            return effData
    else:
        filesToProcess = addendaList + [effFile]
    
    for file in filesToProcess:
        print 'processing file:', file
        # import the data from csv files
        tempDataHolder = {}
        effReader = csv.DictReader(open(file,'rb'),delimiter =',')
        for row in effReader:
            for column,value in row.iteritems():
                tempDataHolder[column] = value
            if re.search('S\d\d\d\d\d',tempDataHolder['Substrate ID']): # filter out any junk data with no substrate ID or data entry mistakes
                substrate = tempDataHolder['Substrate ID'][1:]
                try:
                    if int(substrate)<285:
                        #skip lower substrate numbers for now because running into memoryError
                        continue  
                except:
                    continue
                effData.setdefault(substrate,{})
                effData[substrate].setdefault(tempDataHolder['Web ID'],{})
                for key in colsToImport:
                    effData[substrate][tempDataHolder['Web ID']].setdefault(key,[]).append(tempDataHolder[key])
             
        # add data to sqlite3 database
        tableHeaders = "(\"substrate\" INTEGER, \"web ID\", "
        for col in colsToImport:
            if col in realTypes:
                tableHeaders += "\"" + col + "\" REAL " + ', ' 
            else:
                tableHeaders += "\"" + col + "\"" + ', '  
        tableHeaders = tableHeaders[:-2]  + ')'
        qString = " ?,"*(len(colsToImport) + 2)
        qString = qString[:-1] + ")"
        for substrate in effData.keys():
            cursor.execute("CREATE TABLE IF NOT EXISTS effTable" + tableHeaders)
            for webID in effData[substrate].keys():
                for count in range(len(effData[substrate][webID]['DW'])):
                    insertValues = [substrate, webID] + [effData[substrate][webID][col][count] for col in colsToImport]
                    if file in addendaList:
                        # if appending database with updated data, make sure not adding duplicate data by checking if the DateTested
                        # is already in the database
                        inDB = None
                        try:
                            inDB = cursor.execute("SELECT \"Cell ID\" FROM effTable WHERE \"Cell ID\" = " + 
                                effData[substrate][webID]['Cell ID'][count])
                            cursor.execute("SELECT \"Cell Eff Avg\" FROM effTable WHERE \"Cell Eff Avg\" = " + 
                                effData[substrate][webID]['Cell Eff Avg'][count])
                        except:
                            # if cell already in database, replace with addenda data, because it was probably retested and is now the best test
                            if inDB != None:
                                cursor.execute("DELETE FROM effTable WHERE \"Cell ID\" = " + 
                                effData[substrate][webID]['Cell ID'][count])
                            cursor.execute("INSERT INTO effTable VALUES(" + qString, insertValues)
                    else:
                        try:
                            cursor.execute("INSERT INTO effTable VALUES(" + qString, insertValues)
                        except Exception as e:
                            print e
                            print insertValues
                            raw_input('press enter to continue')
        effDBConn.commit()
        filesProcessed.append(file)

    # re-sort database by substrate and DW if we just added new data
    if len(filesToProcess) > 0:
        tableLabels = "(\"substrate\", \"web ID\", "
        for col in colsToImport:
            if col in realTypes:
                tableLabels += "\"" + col + "\"" + ', ' 
            else:
                tableLabels += "\"" + col + "\"" + ', '  
        tableLabels = tableLabels[:-2]  + ')'
        cursor.execute("CREATE TABLE IF NOT EXISTS temp_effTable" + tableHeaders)
        cursor.execute("INSERT INTO temp_effTable %s SELECT * FROM effTable ORDER BY substrate ASC, DW ASC" % (tableLabels))
        cursor.execute("DROP TABLE effTable")
        cursor.execute("ALTER TABLE temp_effTable RENAME TO effTable")
        effDBConn.commit()
    if effDBConn:
        effDBConn.close()
    if stashFile:
        print 'files processed:',filesProcessed
        pickle.dump(filesProcessed, open(logFile,'wb'))
    
    return effData

def effData_by_substrate(effData = None):
    '''
    Returns a dict of efficiency data with substrate numbers as keys.
    
    :param: effData : dictionary of efficiency data with substrate numbers as primary keys and webIds as secondary keys.
    '''
    if effData == None:
        effData = import_eff_file()
    effDataSubs = {}
    firstRun = effData.keys()[0]
    firstWebID = effData[firstRun].keys()[0]
    effKeys = sorted(effData[firstRun][firstWebID].keys())
    keyCount = 0
    for key in effKeys:
        if key == 'DW':
            dwCount = keyCount + 1 # find the position of DW in effKeys and add one because the first element of the tempEffData list is webID
            break
        keyCount += 1
        
    for run in sorted(effData.keys()):
        tempEffData = []
        effDataSubs.setdefault(run,{})
        for webID in sorted(effData[run].keys()):
            for count in range(len(effData[run][webID]['DW'])):
                tempEffData.append([webID] + [effData[run][webID][key][count] for key in effKeys])
        tempEffData = sorted(tempEffData, key = lambda data : data[dwCount])
        for count in range(len(tempEffData)):
            effDataSubs[run].setdefault('webID',[]).append(tempEffData[count][0])
            keyCount = 1
            for key in effKeys:
                effDataSubs[run].setdefault(key,[]).append(tempEffData[count][keyCount])
                keyCount+=1
                
    return effDataSubs
    
def interp_to_eff(eff_data_DW,dataset_DW,dataset):
    min_effdw=min(eff_data_DW)
    max_effdw=max(eff_data_DW)

    min_eff_dw_index=min(range(len(dataset_DW)), key=lambda i: 
    abs(dataset_DW[i]-min_effdw))
    max_eff_dw_index=min(range(len(dataset_DW)), key=lambda i: 
    abs(dataset_DW[i]-max_effdw))

    dataset_dw_index_min=min(range(len(dataset_DW)), key=lambda i: 
    abs(dataset_DW[i]-min_effdw))
    dataset_dw_index_max=min(range(len(dataset_DW)), key=lambda i: 
    abs(dataset_DW[i]-max_effdw))
    
    if (dataset_dw_index_max-dataset_dw_index_min)>(max_eff_dw_index-min_eff_dw_index):
        dataset_dw_index_max-=1
    if (dataset_dw_index_max-dataset_dw_index_min)<(max_eff_dw_index-min_eff_dw_index):
        dataset_dw_index_max+=1
    
    interped_data=np.interp(eff_data_DW,
    dataset_DW[min_eff_dw_index:max_eff_dw_index],
    dataset[dataset_dw_index_min:dataset_dw_index_max])
    
    return interped_data

def getLatestScheduleFile():
    schedFileDir = 'Y:/ProcessFE/'
    schedFiles = os.listdir(schedFileDir)
    sortedSchedFiles = sorted(schedFiles, key=lambda x: os.stat(os.path.join(schedFileDir, x)).st_mtime)
    schedFile = False
    fileCounter = 0
    while not schedFile:
        fileCounter -= 1
        if re.search('Daily', sortedSchedFiles[fileCounter]) and not re.search('\$', sortedSchedFiles[fileCounter]) and not re.search('~', sortedSchedFiles[fileCounter]):
            latestSchedFile = sortedSchedFiles[fileCounter]
            schedFile = True
            print 'latest schedule file found: ', latestSchedFile
    return schedFileDir + latestSchedFile
    
def getRunDates(stash_dates = True):
    # imports latest efficiency file and grabs the run dates for each Web ID on each tool, writing to the datesFile file.  Also gets DW range for each substrate ID
    # argument stash_dates tells the program to store the processed dates in a csv file or not.
    print 'getting latest run dates...'
    logFile = 'Y:/Nate/get dates of runs/runDateLogFile.txt'
    datesFile = 'Y:/Nate/get dates of runs/all run dates.csv'
    dates = {}
    latestEffFile = getLatestEffFile()
    latestEffFileDate = time.ctime(os.path.getmtime(latestEffFile))
    addendaFiles = get_addenda_eff_files()
    last = 0
    for file in addendaFiles + [latestEffFile]:
        latest = time.ctime(os.path.getmtime(latestEffFile))
        if latest < last:
            last = latest
        
    datesUpToDate = False
    if stash_dates:
        if os.path.isfile(logFile):
            with open(logFile,'a+') as runDateLogFile:
                latestDate = runDateLogFile.readlines()[-1]
                if latestDate == latestEffFileDate:
                    datesUpToDate = True
                else:
                    runDateLogFile.write('latest eff file modified last:\r\n' + last)
                    datesUpToDate = False
                    print 'updating stashed date file with newest data'
        else:
            with open(logFile,'wb') as runDateLogFile:
                runDateLogFile.write('latest eff file modified last:\r\n' + latestEffFileDate)
        
    
    if datesUpToDate:
        print 'using stashed date file'
        dateReader = csv.DictReader(open(datesFile,'rb'),delimiter = ',')
        for row in dateReader:
            dates.setdefault(row['substrate'],{})
            dates[row['substrate']].setdefault(row['Web ID'],{})
            for column,value in row.iteritems():
                if re.search('\d\d/\d\d/\d\d\d\d',value):
                    dates[row['substrate']][row['Web ID']][column] = datetime.datetime.strptime(value, '%m/%d/%Y')
                else:
                    dates[row['substrate']][row['Web ID']][column] = value
    else:
        dates = {}
        dateKeys = ['BC Run','BE Run','SE Run','PC Run','CDS Run','TCO Run']
        toolKeys = ['BE Tool', 'PC Tool']
        recipes = ['BE Recipe', 'BC Recipe', 'PC Recipe', 'Se Recipe', 'TCO Recipe', 'Cds Recipe']
        badKeys =['010101AP', '', '-']
        
        print 'importing efficiency data...'
        effData = import_eff_file()
        print 'finished importing.'
        
        for web in sorted(effData.keys()):
            dates[web] = {}
            
            for webID in sorted(effData[web].keys()):
                dates[web][webID] = {}
                dates[web][webID]['DW start'] = effData[web][webID]['DW'][0]
                dates[web][webID]['DW end'] = effData[web][webID]['DW'][-1]
                
                
                
                for count in range(len(effData[web][webID]['BE Run'])):
                    if reduce(operator.mul, [effData[web][webID][key][count] not in badKeys for key in dateKeys + toolKeys]):
                    
                        for key in toolKeys + recipes + ['Substrate Lot']:
                            dates[web][webID].setdefault(key, effData[web][webID][key][count])
                    
                        for key in dateKeys:
                            tempDate = datetime.datetime.strptime(effData[web][webID][key][count][:6], '%y%m%d')
                            dates[web][webID].setdefault(key, tempDate)
                            
                            if dates[web][webID][key] != tempDate:
                                print web, webID, key
                                print 'date in effData:', effData[web][webID][key][count], 'doesn\'t match date already in dict:', dates[web][webID][key]
                                print 'overwriting...'
                                dates[web][webID][key] = tempDate

            else:
                firstRow = False
                
        print 'writing dates to file'
        with open(datesFile,'wb') as file:
            csvfile = csv.writer(file, delimiter = ',')
            csvfile.writerow(['substrate', 'Web ID', 'DW start', 'DW end'] + toolKeys + dateKeys + recipes + ['Substrate Lot'])
            for web in sorted(dates.keys()):
                for webID in sorted(dates[web].keys()):
                    try:
                        csvfile.writerow([web, webID, dates[web][webID]['DW start'], dates[web][webID]['DW end']] + 
                            [dates[web][webID][key] for key in toolKeys] + 
                            [datetime.datetime.strftime(dates[web][webID][key], '%m/%d/%Y') for key in dateKeys] + 
                            [dates[web][webID][key] for key in recipes] + [dates[web][webID]['Substrate Lot']])
                    except KeyError as ke:
                        print 'key error: ', ke
                        print web, webID, dates[web][webID]
                        
    print 'dates loaded'
            
    return dates

def getOOWls():
    '''
    returns list of wavelengths for nuvosun's usb2000+ ocean optics spectrometers
    '''
    wlReader = csv.reader(open('Y:\Nate\code\oceanOpticsWavelengths200-900.txt','rb'), delimiter = ',')
    wl = [float(x) for x in wlReader.next()]
    return wl

def OESparameters(normalizations = False):
    '''
    Returns a dict with elements as keys and colors, min/max wavelengths.
    
    :param: normalizations : boolean, whether to return normalization keys (/Fi, /Ar-811, /Fi/Ar-811)
    '''
    elementList = ['Cu-325-327','Cu-515','In-451','Ga-417','Se-473','Ar-811','Na-589','Mo-380','Ti-496-522','O-777','H-656','Fi']
    colorList = ['yellow','dark yellow','orange','grey','maroon','red','lightcoral','purple','green','blue','bisque','pink']
    OESminList = [321.0, 513.0, 449.0, 414.0, 470.0, 808.0, 587.0, 378.0, 496.0, 775.0, 654.0, 189.77481] #wavelength minimums for OES integration
    OESmaxList = [330.0, 517.0, 453.0, 418.0, 475.0, 815.0, 589.0, 382.0, 522.0, 779.0, 658.0, 890.3067897] #wavelength maxs

    wl = getOOWls()

    elementDict = {}
    for elementCount in range(len(elementList)):
        elementDict[elementList[elementCount]] = {}
        elementDict[elementList[elementCount]]['color'] = colorList[elementCount]
        elementDict[elementList[elementCount]]['minWL'] = OESminList[elementCount]
        elementDict[elementList[elementCount]]['maxWL'] = OESmaxList[elementCount]
        elementDict[elementList[elementCount]]['minWLindex'],elementDict[elementList[elementCount]]['maxWLindex'] = get_WL_indices(OESminList[elementCount], OESmaxList[elementCount], wl)
    
    # array of keys for normalization of data, used in normalizedData dict
    keysToAdd = [key + '/Fi' for key in elementDict.keys()] + [key + '/Ar-811' for key in elementDict.keys()] + [
        key + '/(Fi*Ar-811)' for key in elementDict.keys()]
    removeKeys = ['Fi/Fi','Ar-811/Ar-811','Fi/Ar-811','Fi/(Fi*Ar-811)','Ar-811/(Fi*Ar-811)']
    for key in removeKeys:
        keysToAdd.remove(key)
    
    if normalizations:
        # array of keys for normalization of data, used in normalizedData dict
        normalizationKeys = [key + '/Fi' for key in elementDict.keys()] + [key + '/Ar-811' for key in elementDict.keys()] + [
            key + '/(Fi*Ar-811)' for key in elementDict.keys()]
        removeKeys = ['Fi/Fi','Ar-811/Ar-811','Fi/Ar-811','Fi/(Fi*Ar-811)','Ar-811/(Fi*Ar-811)']
        for key in removeKeys:
            normalizationKeys.remove(key)
        
        return elementDict, normalizationKeys
    else:
        return elementDict
    
def get_WL_indices(minWL,maxWL,wl):
    #returns the indices of the wl list where the min and max wavelengths are, supplied as minWL and maxWL
    lowerWLindex=min(range(len(wl)), key=lambda i: abs(wl[i]-minWL))
    upperWLindex=min(range(len(wl)), key=lambda i: abs(wl[i]-maxWL))
    return lowerWLindex, upperWLindex
    
def backupFiles(srcDir, destDir):
    #backs up folders and files from source directory srcDir, to destination directory, destDir
    srcFiles = os.listdir(srcDir)
    print 'backing up files from', srcDir, 'to', destDir
    for thing in srcFiles:
        print 'backing up', thing
        if os.path.isfile(srcDir + thing):
            if os.path.isfile(destDir + thing):
                print thing, 'up to date, skipping copy'
            else:
                shutil.copy(srcDir + thing, destDir + thing)
        if os.path.isdir(srcDir + thing):
            if os.path.isdir(destDir + thing):
                print thing, 'up to date, skipping copy'
            else:
                distutils.dir_util.copy_tree(srcDir + thing, destDir + thing)
        
def get_memory_use():
    w = WMI('.')
    result = w.query("SELECT WorkingSet FROM Win32_PerfRawData_PerfProc_Process WHERE IDProcess=%d" % os.getpid())
    return result#['WorkingSet'])

def get_BEPC_excelFiles(runCutoff = 285, lastRun = 500):
    '''
    Returns a list of the full paths to all BE/PC excel files.

    :param: runCutoff : the earliest run to start scanning for
    :param: lastRun : the last run to scan for
    '''
    excelFiles = []
    basepath = 'Y:\\Experiment Summaries\\Year 20'
    years = [13,14,15,16,17]
    runs = range(runCutoff, lastRun)
    for eachRun in runs:
        eachRun = str(eachRun)
        print 'searching for PC excel file for', eachRun
        for year in years:
            runStr = str(eachRun).rjust(5,'0')
            runPath = basepath + str(year) + '\\' + 'S' + runStr + '\\'
            if os.path.exists(runPath):
                foundFile = False
                for f in glob.iglob(runPath + '*S*' + eachRun + '*' + 'MC0[1,2]' + '*xlsx'):
                    if re.search('not complete', f) or re.search('copy', f, re.IGNORECASE) or re.search('~', f) or foundFile:
                        print 'skipping file:'
                        print f
                        continue
                    excelFiles.append(f)
                    foundFile = True

    return excelFiles
    
def getRunDatesFromBEPCexcel():
    '''
    Returns a dictionary of runNumber:[BERunDate,PCRunDate]
    '''
    runDates = {}
    excelFiles = get_BEPC_excelFiles()
    for file in excelFiles:
        run = re.search('S00(\d\d\d)',file).group(1)
        runDates[run] = {}
        print file
        print run
        wb = load_workbook(filename = file, use_iterators=True, data_only=True)
        try:
            BEws = wb.get_sheet_by_name(name = 'BE')
            BEdateStr = BEws['A1'].value
            BEdate = re.search('\d\d\d\d\d\d',BEdateStr).group(0)
            runDates[run]['BE date'] = datetime.datetime.strptime(BEdate, '%y%m%d')
            print run, 'BE run date', datetime.datetime.strptime(BEdate, '%y%m%d')
        except:
            runDates[run]['BE date'] = 'unknown'
        try:
            PCws = wb.get_sheet_by_name(name = 'PC')
            PCdateStr = PCws['A1'].value
            PCdate = re.search('\d\d\d\d\d\d',PCdateStr).group(0)
            runDates[run]['PC date'] = datetime.datetime.strptime(PCdate, '%y%m%d')
            print run, 'PC run date', datetime.datetime.strptime(PCdate, '%y%m%d')
        except:
            runDates[run]['PC date'] = 'unknown'
        runDates[run]['tool'] = re.search('MC\d\d',file).group(0)
        
    
    return runDates
        

def get_XRF_data(runs, minDW = 0):
    '''Returns a dictionary (keys = run numbers) of dictionaries (keys = xrf parameters) by scraping all the excel files.
    
    :param: runs: array of run numbers to be used e.g. [399,400]
    :param: minDW: minimum DW cutoff for grabbing XRF data
    '''
    XRFsheetNames = ["MC01 XRF", "MC02 XRF", "XRF"]
    keys = ['DT','Cu','Ga','Mo','Se','Thickness','In','DW','Cu3','In3','Ga3']
    keysToInterp = ['DT','Cu','Ga','Mo','Se','Thickness','In']
    XRFdata = {}
    runsWithNoXRFfile = []
    tempXRFdata = {}# to use to get raw data, interp to get full data
    basepath = 'Y:\\Experiment Summaries\\Year 20'
    years = [13,14,15,16,17]
    for eachRun in runs:
        eachRun = str(eachRun)
        noXRFfile = True
        print 'searching for XRF file for', eachRun
        for year in years:
            runStr = str(eachRun).rjust(5,'0')
            runPath = basepath + str(year) + '\\' + 'S' + runStr + '\\'
            if os.path.exists(runPath):
                #print runPath
                
                for f in glob.iglob(runPath + '*S*' + eachRun + '*' + 'MC0[1,2]' + '*xlsx'):
                    if re.search('not complete', f) or re.search('copy', f, re.IGNORECASE) or re.search('~', f):
                        print 'skipping file:'
                        print f
                        continue
                    if eachRun == '426':
                        newFormat = True
                    else:
                        newFormat = False
                    if not noXRFfile: # to make sure we only get one XRF file
                        break
                    else:
                        noXRFfile = False
                        
                        xrfFile = f

                        wb = load_workbook(filename = xrfFile, use_iterators=True, data_only=True)  
                        PCwb = wb.get_sheet_by_name(name = 'Web Movement Procedure PC')
                        maxDW = float(PCwb['C6'].value)
                        print 'max DW:',maxDW
                        foundSheet = False
                        for sheetName in XRFsheetNames:
                            try:
                                ws4 = wb.get_sheet_by_name(name = sheetName)
                                foundSheet = True
                            except:
                                pass
                        if not foundSheet:
                            print 'didn\'t find XRF worksheet in file'
                            exit()
                        '''
                        row labels
                        2: Cu
                        3: Ga
                        4: Mo
                        5: Se
                        6: Cu/III
                        7: Thickness
                        8: In
                        9:  blank
                        10: DW
                        '''
                        XRFdata[eachRun]={}
                        tempXRFdata[eachRun]={}
                        for key in keys:
                            XRFdata[eachRun][key]=[]
                            tempXRFdata[eachRun][key]=[]

                            
                        if newFormat: #426 and possible all newer runs have CuIII and Cu columns switched
                            dtRow = 0
                            cuRow = 6
                            gaRow = 3
                            moRow = 4
                            seRow = 5
                            thRow = 7
                            inRow = 8
                            dwRow = 10
                        else:
                            dtRow = 0
                            cuRow = 2
                            gaRow = 3
                            moRow = 4
                            seRow = 5
                            thRow = 7
                            inRow = 8
                            dwRow = 10

                        rowcounter=0
                        for row in ws4.iter_rows():
                            try:
                                if rowcounter==0:
                                    #labels=[row[2].value,row[3].value,row[3].value,row[10].value.encode('utf-8')]
                                    rowcounter+=1
                                    pass
                                elif rowcounter == 1 and row[dwRow].value!= '' and row[dwRow].value!= None and row[dwRow].value>=minDW and row[dwRow].value<=maxDW and float(row[dwRow].value)!=-5.57:
                                # -5.57 shows up in 389 when two rows are missing in 'web length' column # this used to be in there too: and row[dwRow].value<=allrundata[eachRun][1]
                                    if row[dtRow].value!=None:
                                        if type(row[dtRow].value) != datetime.datetime:
                                            try:
                                                tempDate = dateParser(row[dtRow].value)
                                            except Exception as e:
                                                print e
                                                print row[dtRow].value
                                        else:
                                            tempDate = row[dtRow].value
                                        epochDT = (tempDate - datetime.datetime(1970,1,1)).total_seconds()
                                        tempXRFdata[eachRun]['DT'].append([epochDT,row[dwRow].value])
                                    if row[cuRow].value!=None:
                                        tempXRFdata[eachRun]['Cu'].append([row[cuRow].value,row[dwRow].value])
                                    if row[gaRow].value!=None:
                                        tempXRFdata[eachRun]['Ga'].append([row[gaRow].value,row[dwRow].value])
                                    if row[moRow].value!=None:
                                        tempXRFdata[eachRun]['Mo'].append([row[moRow].value,row[dwRow].value])
                                    if row[seRow].value!=None:
                                        tempXRFdata[eachRun]['Se'].append([row[seRow].value,row[dwRow].value])
                                    if row[thRow].value!=None:
                                        tempXRFdata[eachRun]['Thickness'].append([row[thRow].value,row[dwRow].value])
                                    if row[inRow].value!=None:
                                        tempXRFdata[eachRun]['In'].append([row[inRow].value,row[dwRow].value])
                                    XRFdata[eachRun]['DW'].append(row[dwRow].value)
                            except Exception as e:
                                print e
                
        if not noXRFfile: #if found the XRF file
            print 'found xrf file for run ', eachRun
            print xrfFile
            for key in keysToInterp:
                tempXRFdata[eachRun][key]=np.array(tempXRFdata[eachRun][key],dtype='float64')
                XRFdata[eachRun][key]=np.interp(XRFdata[eachRun]['DW'],tempXRFdata[eachRun][key][:,1],tempXRFdata[eachRun][key][:,0])
            for count in range(len(XRFdata[eachRun]['DW'])):
                #print count
                XRFdata[eachRun]['Cu3'].append((XRFdata[eachRun]['Cu'][count])/(XRFdata[eachRun]['In'][count]+XRFdata[eachRun]['Ga'][count]))
                XRFdata[eachRun]['In3'].append((XRFdata[eachRun]['In'][count])/(XRFdata[eachRun]['Cu'][count]+XRFdata[eachRun]['Ga'][count]))
                XRFdata[eachRun]['Ga3'].append((XRFdata[eachRun]['Ga'][count])/(XRFdata[eachRun]['Cu'][count]+XRFdata[eachRun]['In'][count]))
        else:
            print 'didn\'t find XRF file for run', eachRun
            runsWithNoXRFfile.append(eachRun)
            
    return XRFdata
    
def get_saved_runsInOESDBList():
    '''Returns a dict of runs already in the database and a dict of their dates.
    '''
    upToDate = False
    runsInDB = []
    runDatesInDB = []
    lastOESdbMtime = None
    if os.path.isfile(OESmtimeFile) and os.path.isfile(OESdbFile):
        lastOESdbMtime = os.path.getmtime(OESdbFile)
        lastLoggedTime = pickle.load(open(OESmtimeFile))
        print '########OESDBmtimes: ', lastLoggedTime, lastOESdbMtime
        if lastLoggedTime == lastOESdbMtime:
            if os.path.isfile(runsInDBFile) and os.path.isfile(runDatesInDBFile):
                runsInDB = pickle.load(open(runsInDBFile))
                runDatesInDB = pickle.load(open(runDatesInDBFile))
                upToDate = True
                print 'using stashed files for list of OES files in DB'
    return upToDate, lastOESdbMtime, runsInDB, runDatesInDB
    
def set_saved_runsInOESDBList(runsInDB, runDatesInDB):
    '''Returns nothing, saves pickle files of runs and run dates already in the OES database.
    
    :param: runsInDB: list of runs that have been entered in the database.
    :param: runDateLogFile: list of run dates that have been entered in database (folders with data are labeled by date, not run)
    '''
    OESmtimeFile = 'Y:/Nate/OES/records/OESmtime.pkl'
    runsInDBFile = 'Y:/Nate/OES/records/runsInDB.pkl'
    runDatesInDBFile = 'Y:/Nate/OES/records/runDatesInDB.pkl'
    OESdbFile = 'Y:/Nate/OES/databases/all OES data.csv'
    pickle.dump(os.path.getmtime(OESdbFile), open(OESmtimeFile, 'wb'))
    pickle.dump(runsInDB, open(runsInDBFile, 'wb'))
    pickle.dump(runDatesInDB, open(runDatesInDBFile, 'wb'))
    return
    
def get_saved_runsInprocessedOESDBList():
    '''Returns two booleans that tell if the processed data is up to date with the raw data database,
    and dicts of PC BE runs already in the processed databases.
    
    '''
    writtenPCrunsFile = 'Y:/Nate/OES/records/processed_OESPCruns.pkl'
    writtenBErunsFile = 'Y:/Nate/OES/records/processed_OESBEruns.pkl'
    OESDTDWFile = 'Y:/Nate/OES/records/processed_OESDTDW.pkl'
    if os.path.isfile(writtenPCrunsFile) and os.path.isfile(writtenPCrunsFile) and os.path.isfile(OESDTDWFile):
        with open(writtenPCrunsFile) as PCfile:
            writtenPCruns = pickle.load(PCfile)
        with open(writtenBErunsFile) as BEfile:
            writtenBEruns = pickle.load(BEfile)
        with open(OESDTDWFile) as dtfile:
            OESDTDW = pickle.load(dtfile)
    else:
        writtenPCruns = []
        writtenBEruns = []
        OESDTDW = None
    
    processedPCUpToDate = True
    processedBEUpToDate = True
    processed_OESmtimeFile = 'Y:/Nate/OES/records/PROCESSED_OESmtime.pkl'
    processed_runsInDBFile = 'Y:/Nate/OES/records/PROCESSED_runsInDB.pkl'
    processed_runDatesInDBFile = 'Y:/Nate/OES/records/PROCESSED_runDatesInDB.pkl'
    
    upToDate = False
    runsInDB = []
    runDatesInDB = []
    lastOESdbMtime = None
    if os.path.isfile(processed_OESmtimeFile) and os.path.isfile(processed_OESdbFile):
        lastOESdbMtime = os.path.getmtime(processed_OESdbFile)
        lastLoggedTime = pickle.load(open(processed_OESmtimeFile))
        print '########processed_OESDBmtimes: ', lastLoggedTime, lastOESdbMtime
        if lastLoggedTime == lastOESdbMtime:
            if os.path.isfile(processed_runsInDBFile) and os.path.isfile(processed_runDatesInDBFile):
                runsInDB = pickle.load(open(processed_runsInDBFile))
                runDatesInDB = pickle.load(open(processed_runDatesInDBFile))
                upToDate = True
                print 'using stashed files for list of OES files in DB'
    
    for run in writtenBEruns:
        if run in runsInDB:
            pass
        else:
            processedBEUpToDate = False
    if writtenBEruns == []:
        processedBEUpToDate = False
    for run in writtenPCruns:
        if run in runsInDB:
            pass
        else:
            processedPCUpToDate = False
    if writtenPCruns == []:
        processedPCUpToDate = False
    
    return processedPCUpToDate, processedBEUpToDate, writtenPCruns, writtenBEruns, OESDTDW
    
def set_saved_runsInprocessedOESDBList(PCruns, BEruns, OESDTDW):
    '''Returns nothing, saves pickle files of runs and run dates already in the processed OES database.
    
    :param: PCruns: list of PC runs that have been entered in the PROCESSED OES database.
    :param: BEruns: list of BE runs that have been entered in the PROCESSED OES database.
    '''
    writtenPCrunsFile = 'Y:/Nate/OES/records/processed_OESPCruns.pkl'
    writtenBErunsFile = 'Y:/Nate/OES/records/processed_OESBEruns.pkl'
    OESDTDWFile = 'Y:/Nate/OES/records/processed_OESDTDW.pkl'
    if os.path.isfile(writtenPCrunsFile) and os.path.isfile(writtenPCrunsFile):
        with open(writtenPCrunsFile) as PCfile:
            writtenPCruns = pickle.load(PCfile)
        with open(writtenBErunsFile) as BEfile:
            writtenBEruns = pickle.load(BEfile)
    else:
        writtenPCruns = []
        writtenBEruns = []
    
    with open(writtenPCrunsFile,'wb') as PCfile:
        try:
            pickle.dump(PCruns + writtenPCruns, PCfile)
        except NameError:
            pickle.dump(PCruns, PCfile)
    with open(writtenBErunsFile,'wb') as BEfile:
        try:
            pickle.dump(BEruns + writtenBEruns, BEfile)
        except NameError:
            pickle.dump(BEruns, BEfile)
    with open(OESDTDWFile, 'wb') as dwfile:
        pickle.dump(OESDTDW, dwfile)

    return