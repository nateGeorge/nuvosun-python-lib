from openpyxl import load_workbook
import glob, os, numpy, re, datetime, pyodbc, csv, struct, matplotlib, time, sys, easygui
from scipy import signal
import pylab as plt
from mpl_toolkits.mplot3d import Axes3D
import matplotlib.gridspec as gridspec
from matplotlib import cm
from scipy import interpolate
import Tkinter as tk
import tkFileDialog as tkFD
import glob


#gets min/max datetime and DW position for specified MC03 runs
#*************************************************
'''runstart=370
runend=375
runlist=range(runend-runstart+1)
runlist=[x+runstart for x in runlist]
runlist.remove(353)
runlist.remove(339)
runlist.remove(340)
runlist.remove(343)
runlist.remove(357)
runlist.remove(358)
runlist.remove(361)'''
#runlist=[382]#, 338, 349, 350, 354, 355, 364, 365]

runnumber=raw_input('enter run number (i.e. 365)')

runlist=[runnumber]

try:
	todaysdate=input('for previous data, enter date as YYMMDD (otherwise defaults to today): ')
	todaysdate=str(todaysdate)
except SyntaxError: # if a date is not entered, use today's date
	todaysdate=time.strftime("%y%m%d")

webspeed=0.684
if re.search('R',runnumber,re.IGNORECASE) or re.search('L',runnumber,re.IGNORECASE): # slower webspeed for 13" runs
	webspeed=0.380

	
rootdir = "Y:\Experiment Summaries\Year 20"+todaysdate[:2]+"\\"
scanDir = rootdir+"S00" + str(runnumber[:3])+'\\'

# redirect standard error and standard out to log files
sys.stderr = open(scanDir + '/MC03 OES/' + 'error.log', 'w')
sys.stdout = open(scanDir + '/MC03 OES/' + 'output.log', 'w')

'''if not os.path.isdir(scanDir):
        print 'couldn\'t find directory : ', scandir
	lastyear = str(int(todaysdate[:2])-1)
	scanDir = "Y:\Experiment Summaries\Year 20" + lastyear + "\\" + "S00" + str(runnumber[:3])+'\\'
	if not os.path.isdir(scanDir):
                print 'also couldn\'t find directory : ', scandir
		print 'choose the directory to scan for the file'
		root = tk.Tk()
		root.withdraw()
		root.wm_attributes("-topmost",1)
		scanDir = tkFD.askdirectory(parent=root,initialdir='Y:\Experiment Summaries\\')
'''


for IRcable in [0,1]:
	if IRcable == 0:
		dwoffset = 6.57#1.51 + 5.06
		cable = 0
		timeoffset = dwoffset/webspeed*60
	elif IRcable == 1:
		dwoffset = 0
		cable = 7 #the multiplexer number is actually 7 for takeup IR sensor
		timeoffset = 0
	rundates=[]
	runnums=[]
	runfilenames=[]
	DTmins=[]
	DTmaxs=[]
	DWmins=[]
	DWmaxs=[]
	DWstarts=[]
	DWends=[]
	DWexactstarts=[]
	DWexactends=[]

	for i in runlist:
		currentrun=i
		subruncount=0
		for f in glob.iglob(scanDir + 'MC03*Checklist*' + todaysdate + '*.xlsx'):
			print 'using ', f, 'as file for DW positions'
			if subruncount==0:
				runfilenames.append(f)
				rundates.append(re.search('\d\d-*\d\d-*\d\d',f).group(0))
				dateFromFilename = re.search('\d\d-*\d\d-*\d\d',f).group(0)
				runnums.append(currentrun)
			if subruncount>0:
				runfilenames.append(f)
				rundates.append(re.search('\d\d-*\d\d-*\d\d',f).group(0))
				runnums.append(str(currentrun)+'-'+str(subruncount))
			subruncount+=1
			
			if re.search('-\s*[DW]*\s*(\d+)\s*(to|-+)\s*(\d+)',f).group(1):
				DWstart=float(re.search('-\s*[DW]*\s*(\d+)\s*(to|-+)\s*(\d+)',f).group(1))
			else:
				DWstart=raw_input('enter DW start: ')
			if re.search('-\s*[DW]*\s*(\d+)\s*(to|-+)\s*(\d+)',f).group(3):
				DWend=float(re.search('-\s*[DW]*\s*(\d+)\s*(to|-+)\s*(\d+)',f).group(3))
			else:
				DWend=raw_input('enter DW end: ')
			print 'detected DWstart: ', DWstart
			print 'detected DWend: ', DWend
			if DWstart>DWend:
				DWmax=DWstart
				DWmin=DWend
			else:
				DWmax=DWend
				DWmin=DWstart
			
			wb = load_workbook(filename = f,use_iterators=True, data_only=True)	
			ws4 = wb.get_sheet_by_name(name="Absorption & Reflection")
			
			dates=[]
			DWpos=[]
			for row in ws4.iter_rows(): #get the start/end datetimes from absorbance data in excel report
				if row[0].value!=None and isinstance(row[0].value, datetime.datetime) and row[5].value!=None and row[5].value<=(DWmax-1.51) and row[5].value>=(DWmin-1.51):
					DWpos.append(row[5].value)
					dates.append(row[0].value)
			
			print 'datetime start from excel file: ', dates[0]
			print 'datetime end from excel file: ', dates[-1]
			dateInFile = datetime.datetime.strftime(dates[-1],'%y%m%d')
			if dateInFile != dateFromFilename:
				print 'date in file ' + dateInFile + ' did not match date in filename ' + dateFromFilename
				currentdir="Y:\Experiment Summaries\MC03\\"
				foundOtherFile = False
				for f in glob.iglob(currentdir + 'MC03 Checklist-'+todaysdate+'*.xlsx'):
					print 'using ' + f + ' instead'
					foundOtherFile = True
				if not foundOtherFile:
					easygui.msgbox('couldn\'t find the excel file, make sure it is named properly, is in the right folder, and has the absorption data with DW in it\n\nbless up,\n -Nate')
					exit()
				if re.search('-[DW]*\s*(\d+)\s*(to|-+)\s*(\d+)',f).group(1):
					DWstart=float(re.search('-[DW]*\s*(\d+)\s*(to|-+)\s*(\d+)',f).group(1))
				else:
					DWstart=raw_input('enter DW start: ')
				if re.search('-[DW]*\s*(\d+)\s*(to|-+)\s*(\d+)',f).group(3):
					DWend=float(re.search('-[DW]*\s*(\d+)\s*(to|-+)\s*(\d+)',f).group(3))
				else:
					DWend=raw_input('enter DW end: ')
				print 'detected DWstart: ', DWstart
				print 'detected DWend: ', DWend
				if DWstart>DWend:
					DWmax=DWstart
					DWmin=DWend
				else:
					DWmax=DWend
					DWmin=DWstart
				
				wb = load_workbook(filename = f,use_iterators=True, data_only=True)	
				ws4 = wb.get_sheet_by_name(name="Absorption & Reflection")
				
				dates=[]
				DWpos=[]
				for row in ws4.iter_rows(): #get the start/end datetimes from absorbance data in excel report
					if row[0].value!=None and isinstance(row[0].value, datetime.datetime) and row[5].value!=None and row[5].value<=(DWmax-1.51) and row[5].value>=(DWmin-1.51):
						DWpos.append(row[5].value)
						dates.append(row[0].value)
					
			DTmins.append(dates[0])
			DTmaxs.append(dates[-1])
			DWmins.append(DWmin)
			DWmaxs.append(DWmax)
			DWstarts.append(DWstart)
			DWends.append(DWend)
			DWexactstarts.append(DWpos[0])
			DWexactends.append(DWpos[-1])
	alldata=numpy.vstack((runnums,DTmins))
	alldata=numpy.vstack((alldata,DTmaxs))
	alldata=numpy.vstack((alldata,DWstarts))
	alldata=numpy.vstack((alldata,DWends))
	alldata=numpy.vstack((alldata,DWexactstarts))
	alldata=numpy.vstack((alldata,DWexactends))
	alldata=alldata.transpose()

	#0=runnum
	#1=DTmin
	#2=DTmax
	#3=DWstart
	#4=DWend
	#5=DWstart lined up with time
	'''
	#combines the run info data into one array		
	runinfo=zip(runnums, rundates, runfilenames)

	# sorts that array by date
	sorted_runinfo=sorted(runinfo, key=lambda runinfo: runinfo[1])
	print sorted_runinfo
	exit()
	sorted_run_nums=[x[0] for x in sorted_runinfo]

	#gets start/end DW and datetimes for each run
	for run in sorted_run_nums:
		currentdir=rootdir+"S00" + str(run)+'\\'
		for f in glob.iglob(currentdir + "MC03*Checklist*.xlsx"):
			DWstart=float(re.search('-[DW]*\s*(\d+)\s*(to|-)\s*(\d+)',f).group(1))
			DWend=float(re.search('-[DW]*\s*(\d+)\s*(to|-)\s*(\d+)',f).group(3))
			if DWstart>DWend:
				DWmax=DWstart
				DWmin=DWend
			else:
				DWmax=DWend
				DWmin=DWstart
			
			wb = load_workbook(filename = f,use_iterators=True, data_only=True)	
			ws4 = wb.get_sheet_by_name(name="Absorption & Reflection")

			dates=[]
			DWpos=[]
			print run
			print f
			for row in ws4.iter_rows(): #get the start/end datetimes from absorbance data in excel report
				if row[0].value!=None and isinstance(row[0].value, datetime.datetime) and row[5].value!=None and row[5].value<=DWmax and row[5].value>=DWmin:
					DWpos.append(row[5].value)
					dates.append(row[0].value)
			DTmin=dates[0]
			DTmax=dates[-1]
			DTmins.append(DTmin)
			DTmaxs.append(DTmax)
			DWmins.append(DWmin)
			DWmaxs.append(DWmax)
	alldata=numpy.vstack((sorted_run_nums,DTmins))
	alldata=numpy.vstack((alldata,DTmaxs))
	alldata=numpy.vstack((alldata,DWstart))
	alldata=numpy.vstack((alldata,DWend))
	alldata=alldata.transpose()'''



	#****************************

	#gets the IR reflection data

	for entry in alldata:
		#import spectral data
		#set up run number and filename
		rundate=entry[1].strftime('%y%m%d')
		runnum=entry[0]
		MDB = 'Y:\Nate\MC03 optics program\SPEC\SP'+rundate+'.mdb'; DRV = '{Microsoft Access Driver (*.mdb)}'; PWD = ''
		print 'using ', MDB, ' as database file'
		print 'using ', 'Y:\Nate\MC03 optics program\SPEC\SP'+rundate+'.txt', ' as spectra file'
		

		# connect to database
		con = pyodbc.connect('DRIVER={};DBQ={};PWD={}'.format(DRV,MDB,PWD))
		cur = con.cursor()

		# run a query and get all the data from the IR datafile
		SQL = 'SELECT DT, V3, START, LENGTH, WLX0, WLX1, WLX2, WLX3, WLX4 FROM SPECLOG;'
		rows = cur.execute(SQL).fetchall()
		cur.close()
		con.close()
		data=numpy.array(rows)
		counter=0
		
		times=[]

		for each in data:
			if each[1]==str(cable) and (each[0]-entry[1]).total_seconds()>timeoffset and (entry[2]-each[0]).total_seconds()>timeoffset:
				with open('Y:\Nate\MC03 optics program\SPEC\SP'+rundate+'.txt', "rb",1) as f:
					f.seek(each[2],0) # goes to the start of the raw data from the mdb file
					bytesread=f.read(each[3])
					X=[]
					Y=[]
					for i in range(each[3]/2-1): # uses length from mdb file
						Y.append(struct.unpack_from('H', bytesread,i*2)[0])
						X.append(each[4]+each[5]*i+each[6]*i**2+each[7]*i**3+each[8]*i**4)
				times.append(counter)
				counter+=1
				Xtemp=X[31:190] #trims data to 1100 nm to 1650 nm range, consistent with TCO monitor program and best view of the data
				Ytemp=Y[31:190]
				#Y_avg+=Xtemp
				if counter>1: #if not first time through loop
					#all_X_data=numpy.vstack((all_X_data,Xtemp))
					#print Xtemp.shape, Ytemp.shape
					f = interpolate.interp1d(Xtemp, Ytemp, kind='cubic', axis=0)
					Ynew=f(X1_data)
					all_Y_data=numpy.vstack((all_Y_data,numpy.array(Ynew)))
					'''maxs_index_now=signal.find_peaks_cwt(Ytemp,numpy.arange(10,50))
					maxs_indices=numpy.hstack((maxs_indices,maxs_index_now[-1]))
					maxs=numpy.vstack((maxs,[Xtemp[maxs_index_now[-1]],Ytemp[maxs_index_now[-1]]]))
					maxs_time_index=numpy.hstack((maxs_time_index,counter-1))'''
				else:
					'''maxs_index=signal.find_peaks_cwt(Ytemp,numpy.arange(10,50))
					maxs_indices=maxs_index[-1]
					maxs_time_index=counter-1
					maxs=[Xtemp[maxs_index[-1]],Ytemp[maxs_index[-1]]]'''
					X1_data=Xtemp
					all_Y_data=numpy.array(Ytemp)
		if counter==0:
			print "check the last entries in the absorption part of the spreadsheet"
			print "it looks like they are wrong."
			print "exiting program"
			time.sleep(10)
			exit()
		
		X1_data=numpy.array(X1_data,dtype='float')
		all_Y_data2=all_Y_data/1000
		all_Y_data=numpy.transpose(all_Y_data)/1000.0
		X1_data2=numpy.transpose(X1_data)
		down_web_pos=plt.arange((entry[3]),(entry[4]),((entry[4]-entry[3])/len(times)))
		if len(down_web_pos) > all_Y_data.shape[1]:
			down_web_pos=numpy.delete(down_web_pos,-1)
		'''maxsdw=numpy.empty(len(maxs_time_index))
		indexcounter=0
		for index in maxs_time_index:
			maxsdw[indexcounter]=down_web_pos[index]
			indexcounter+=1'''
		with open(scanDir+'/MC03 OES/'+str(entry[0])+'_DW'+str(entry[3])+'-'+str(entry[4])+'-IR'+str(IRcable)+'.csv', 'wb') as fp:
			thematrix = csv.writer(fp, delimiter=',')
			thematrix.writerow(numpy.hstack((['wavelength (down)/DW (sideways)'],down_web_pos)))
			thematrix.writerows(numpy.transpose(numpy.vstack((X1_data,all_Y_data2))))
		
		
		
		
		#plotting
		
		font = {'weight' : 'bold',
				'size'   : 22}

		matplotlib.rc('font', **font)
		
		
		

		
		range_for_Yaxis_scale = [int(round((all_Y_data.shape[1]*0.25))), int(round((all_Y_data.shape[1]*0.75)))] # scales data so that max is only from middle 50% of data
		#print abs(all_Y_data[:, range_for_Yaxis_scale[0]:range_for_Yaxis_scale[1]]).max()
		#print down_web_pos[range_for_Yaxis_scale[0]], down_web_pos[range_for_Yaxis_scale[1]]
		
		#cmap = plt.cm.PRGn
		
		#exit()
		fig=plt.figure(figsize=(20,15),facecolor='w',tight_layout=None)
		ax = fig.gca(projection='3d')
		cont_x,cont_y=plt.meshgrid(down_web_pos,X1_data)
		
		surf = ax.plot_surface(cont_x,cont_y,all_Y_data, rstride=1, cstride=1, cmap=cm.jet,
			linewidth=0, antialiased=False, vmax = abs(all_Y_data[:, range_for_Yaxis_scale[0]:range_for_Yaxis_scale[1]]).max())
		#plt.cm.colors.Normalize(vmax=, vmin=0)# -abs(all_Y_data).max())
		'''scatx,scaty=plt.meshgrid(maxsdw,maxs[:,0])
		ax.scatter(maxsdw,maxs[:,0],maxs[:,1])'''
		fig.colorbar(surf, shrink=0.5, aspect=5, pad=-0.1)
		#ax.view_init(elev=60, azim=-20)
		ax.view_init(elev=90, azim=0)
		ax.w_zaxis.line.set_lw(0.)
		ax.set_zticks([])
		ax.set_ylim((1100,1650))
		plt.xlabel('down web position')
		plt.ylabel('\nwavelength')
		#figManager = plt.get_current_fig_manager()
		#figManager.window.showMaximized()
		
		fig.savefig(scanDir+'/MC03 OES/'+str(entry[0])+'_DW'+str(entry[3])+'-'+str(entry[4])+'-IR'+str(IRcable)+'.png',bbox_inches='tight')
		plt.close()
	
		'''for curvecounter in range(len(X1_data)):
			print X1_data
			print all_Y_data[curvecounter]
			plt.plot(X1_data,all_Y_data[curvecounter])
			#plt.scatter(maxs[curvecounter,0],maxs[curvecounter,1])
			plt.show()'''

if subruncount == 0:
        easygui.msgbox('couldn\'t find excel file for run ' + runnumber + ' with date ' + todaysdate + '...double check your date and run number', ok_button = 'well, shoot')
